from modules.process_data import update_cpfs_in_excel, cross_cpfs, union_data
from modules.email_functions import no_new_subscribers, new_subscribers
from modules.logger import get_logger
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import pandas as pd

# Global logger
logger = get_logger()

class ExcelProcessor:
    """Class to manage Excel file operations, including loading, updating, and data comparison."""

    def __init__(self, excel_path: Path, output_path: Path):
        self.excel_path = excel_path
        self.output_path = output_path
        self.workbook = None
        self.sheet1 = None
        self.sheet2 = None

    def load_workbook(self) -> None:
        """Loads the Excel file and stores the necessary sheets."""
        if not self.excel_path.exists():
            logger.error(f"File {self.excel_path} not found!")
            raise FileNotFoundError(f"File {self.excel_path} not found!")

        try:
            self.workbook = load_workbook(self.excel_path)
            self.sheet1 = self.workbook.active  # First sheet
            self.sheet2 = self.workbook.worksheets[1]  # Second sheet for comparison
        except Exception as e:
            logger.exception(f"Error loading Excel file {self.excel_path}: {e}")
            raise

    def update_and_cross_cpfs(self, df: pd.DataFrame) -> str:
        """Updates CPFs in the Excel file and cross-checks them to find new subscribers."""
        update_cpfs_in_excel(self.sheet1, self.workbook, self.excel_path, df)
        return cross_cpfs(self.sheet1, self.sheet2, self.excel_path, self.workbook, self.output_path)

def get_file_paths() -> tuple[Path, Path]:
    """Returns the paths for the base Excel file and the output file."""
    date_str = datetime.now().strftime("%d%m%Y")
    base_path = Path(__file__).parent

    excel_file = base_path / '05 - farmacia.xlsx'
    output_file = base_path / f'19013906000179 - DR_CENTRAL_FARMACIAS {date_str}.xlsx'
    
    return excel_file, output_file

def main() -> None:
    """Main function to process the Excel data and send an email with new subscribers."""
    excel_file, output_file = get_file_paths()
    
    processor = ExcelProcessor(excel_file, output_file)
    processor.load_workbook()

    df = union_data()
    response = processor.update_and_cross_cpfs(df)

    if response != 'EMAIL ENVIADO':
        new_subscribers(output_file)

if __name__ == "__main__":
    main()

