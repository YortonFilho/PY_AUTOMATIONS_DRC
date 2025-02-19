from sql_queries.sql import SQL_DEPENDENTS_CPFS, SQL_OIMARK_CPFS, SQL_EYAL_CPFS, SQL_COLLABORATORS_CPFS
from modules.email_functions import no_new_subscribers
from modules.oracle_database import Oracle
from modules.logger import get_logger
from datetime import datetime
import pandas as pd
import openpyxl

# Logger configuration
logger = get_logger()

# date variables
date = datetime.now()

def clean_cpf(cpf: str) -> str:
    """
    Remove non numeric characters and return a string
    """
    try:
        # Check if cpf is a string
        if isinstance(cpf, str):
            # Filter numeric characters
            return ''.join(filter(str.isdigit, cpf))
        return ''
    except Exception as e:
        logger.error(f"Erro ao tratar CPFs! {e}")

def update_cpfs_in_excel(
        sheet1: openpyxl.worksheet.worksheet.Worksheet, 
        workbook: openpyxl.workbook.workbook.Workbook, 
        excel_file: str, 
        df: pd.DataFrame
    ) -> None:

    """
    Function to clear spreedsheet and insert collected CPFs into the spreadsheet

    :param sheet1: The first sheet of the spreadsheet
    :param workbook: The workbook object
    :param excel_file: The Excel file
    :param df: DataFrame with CPFs from the database
    """
    try:
        # Clear the first sheet while preserving the header
        for row in range(2, sheet1.max_row + 1):  # Start at the second row
            for col in range(1, 3):  # Clear columns 1 and 2
                sheet1.cell(row=row, column=col).value = None

        # Add header if necessary
        if sheet1['A1'].value is None:
            sheet1['A1'] = "Nome"
        if sheet1['B1'].value is None:
            sheet1['B1'] = "CPF"

        # Create the list to store the formatted data
        rows = [(row.NOME, clean_cpf(str(row.CPF))) for row in df.itertuples(index=False)]

        # Find the first empty row in the first columns
        next_empty_row = 2
        while sheet1.cell(row=next_empty_row, column=1).value is not None:
            next_empty_row += 1

        # add the data to the Excel file in the first two columns only
        for name, cpf in rows:
            # Insert data in columns 1 and 2
            sheet1.cell(row=next_empty_row, column=1, value=name)
            sheet1.cell(row=next_empty_row, column=2, value=cpf)
            next_empty_row += 1  # Move to the next empty row

        # Save changes to the spreadsheet
        try:
            workbook.save(excel_file)
        except Exception as e:
            logger.info(f"Erro ao salvar planilha {excel_file}: {e}!")
            raise e
        
    except Exception as e:
        logger.info(f"Erro ao inserir CPFs à planilha {excel_file}: {e}")
        raise e

def cross_cpfs(
        sheet1: openpyxl.worksheet.worksheet.Worksheet, 
        sheet2: openpyxl.worksheet.worksheet.Worksheet, 
        excel_file: str, 
        workbook: openpyxl.workbook.workbook.Workbook, 
        output_file: str
    ):

    """
    Function to cross all cpfs in two sheets to return only the new cpfs

    :param sheet1: First sheet of the spreadsheet
    :param sheet2: Second sheet of the spreadsheet
    :param workbook: The workbook object
    :param excel_file: The Excel file
    :param output_file: Output file
    """
    try:
        # Collect CPFs from the second sheet
        cpfs_sheet2 = {clean_cpf(sheet2.cell(row=row, column=1).value) 
            for row in range(2, sheet2.max_row + 1) 
            if isinstance(sheet2.cell(row=row, column=1).value, str)}

        # Identify the CPFs that are not in the second sheet
        new_cpfs = [(sheet1.cell(row=row, column=1).value, sheet1.cell(row=row, column=2).value)
            for row in range(2, sheet1.max_row + 1) 
            if clean_cpf(sheet1.cell(row=row, column=2).value) not in cpfs_sheet2]

        # Check if any CPFs were found
        if not new_cpfs:
            logger.info("Todos os CPFs da primeira aba já foram enviados para a segunda aba!")
            # Function to send email saying there are no new subscriptions
            no_new_subscribers()
            return 'EMAIL ENVIADO'
        else:
            # Create a new spreadsheet for the CPFs not found
            df_new_cpfs = pd.DataFrame(new_cpfs, columns=["Nome", "CPF"])
            df_new_cpfs.to_excel(output_file, index=False)
            logger.info(f"Planilha com os novos CPFs criada: {output_file}")

            # Adding the new CPFs to the second sheet with the modification date
            next_empty_row_sheet2 = sheet2.max_row + 1  # Next empty row in the second sheet
            for name, cpf in new_cpfs:
                sheet2.cell(row=next_empty_row_sheet2, column=1, value=cpf)  # CPF in column 1
                sheet2.cell(row=next_empty_row_sheet2, column=2, value=date.strftime('%d/%m/%Y')) 
                next_empty_row_sheet2 += 1

            # Saving the changes to the second sheet
            try:
                workbook.save(excel_file)
                logger.info("Novos CPFs adicionados à segunda aba com sucesso!")
            except Exception as e:
                logger.info(f"Erro ao salvar as alterações na segunda aba: {e}")
                raise e
            
    except Exception as e:
        logger.info(f"Erro ao cruzar CPFs entre as abas da planilha {excel_file}: {e}")
        raise e
    
def union_data() -> pd.DataFrame:
    """
    Function to extract and union the collected CPFs from sqls
    """
    # Initialize class to use functions
    oracle = Oracle()
    
    # List of SQL queries to extract CPFs from different tables
    sql_queries = [
        SQL_DEPENDENTS_CPFS, 
        SQL_OIMARK_CPFS, 
        SQL_EYAL_CPFS, 
        SQL_COLLABORATORS_CPFS
    ]
    
    # Initialize an empty list to hold dataframes
    dfs = []
    
    # Loop over each SQL query, extract the data, and create dataframes
    for sql in sql_queries:
        data = oracle.extract_data(sql=sql)
        df = pd.DataFrame(data, columns=['NOME', 'CPF'])
        dfs.append(df)
    
    # Concatenate all dataframes into one
    df = pd.concat(dfs, ignore_index=True)
    
    return df
